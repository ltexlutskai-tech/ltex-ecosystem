"""
Parse L-TEX Excel files and generate JSON seed data.
Run: python3 packages/db/prisma/parse-excel.py
"""
import openpyxl
import json
import re
import os

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
DATA_OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

# ─── Helpers ─────────────────────────────────────────────────────────────────

TRANSLIT = {
    'а':'a','б':'b','в':'v','г':'h','ґ':'g','д':'d','е':'e','є':'ye',
    'ж':'zh','з':'z','и':'y','і':'i','ї':'yi','й':'y','к':'k','л':'l',
    'м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t','у':'u',
    'ф':'f','х':'kh','ц':'ts','ч':'ch','ш':'sh','щ':'shch','ь':'',
    'ю':'yu','я':'ya',
}

def slugify(text):
    result = ''.join(TRANSLIT.get(c, c) for c in text.lower())
    result = re.sub(r'[^a-z0-9]+', '-', result)
    return result.strip('-')

def parse_quality(text):
    text_lower = text.lower()
    if 'екстра' in text_lower: return 'extra'
    if 'крем' in text_lower: return 'cream'
    if '1й сорт' in text_lower or '1-й' in text_lower: return 'first'
    if '2й сорт' in text_lower or '2-й' in text_lower: return 'second'
    if 'сток' in text_lower: return 'stock'
    return 'mix'

def parse_season(text):
    text_lower = text.lower()
    if 'зима' in text_lower: return 'winter'
    if 'літо' in text_lower: return 'summer'
    if 'демісезон' in text_lower: return 'demiseason'
    return ''

def parse_name_field(name_field):
    """Parse: 'Назва товару (1234), https://youtube..., 25' """
    parts = str(name_field).split(', ')

    name = parts[0].strip()
    code = ''
    video_url = None
    avg_weight = None

    # Extract code from name: (1234)
    code_match = re.search(r'\((\d{4})\)', name)
    if code_match:
        code = code_match.group(1)

    for part in parts[1:]:
        part = part.strip()
        if 'youtube' in part or 'youtu.be' in part:
            video_url = part
        else:
            try:
                w = float(part.replace(',', '.').replace('-', '').strip())
                if 0 < w < 100:
                    avg_weight = w
            except (ValueError, AttributeError):
                pass

    return name, code, video_url, avg_weight

# ─── Category detection ─────────────────────────────────────────────────────

CATEGORY_RULES = [
    # (keywords in name, category_slug, subcategory_slug)
    (['кросівки', 'кроси'], 'vzuttia', 'krosivky'),
    (['черевики', 'черевик'], 'vzuttia', 'cherevyky'),
    (['чоботи', 'чобіт'], 'vzuttia', 'choboty'),
    (['туфлі'], 'vzuttia', 'tufli'),
    (['сандалі', 'сандал'], 'vzuttia', 'sandali'),
    (['шльопанці', "в'єтнамки", 'крокси'], 'vzuttia', 'shlopantsi'),
    (['взуття'], 'vzuttia', 'inshe-vzuttia'),
    (['сумки', 'ремні', 'рюкзак'], 'aksesuary', 'sumky'),
    (['постіль', 'ковдр', 'подушк'], 'dim-ta-pobut', 'postil'),
    (['штори'], 'dim-ta-pobut', 'shtory'),
    (['рушник'], 'dim-ta-pobut', 'rushnyky'),
    (['іграшк', 'м\'яких'], 'igrashky', 'miaki'),
    (['bric', 'брік', 'bric-a-brac'], 'bric-a-brac', 'miks-bric'),
    (['косметик'], 'kosmetyka', 'miks-kosmetyka'),
    (['agd', 'товари для'], 'dim-ta-pobut', 'inshe-dim'),
    (['футболк'], 'odyag', 'futbolky'),
    (['сорочк'], 'odyag', 'sorochky'),
    (['світшот', 'кофт', 'свитшот', 'худі'], 'odyag', 'svitshoty'),
    (['толстовк'], 'odyag', 'tolstovky'),
    (['светр', 'свитер'], 'odyag', 'svetry'),
    (['куртк', 'анорак', 'anorak', 'вітровк', 'пуховик'], 'odyag', 'kurtky'),
    (['пальто'], 'odyag', 'palto'),
    (['жилет', 'безрукав'], 'odyag', 'zhylety'),
    (['джинс'], 'odyag', 'dzhinsy'),
    (['штан', 'брюк', 'карго'], 'odyag', 'shtany'),
    (['шорт'], 'odyag', 'shorty'),
    (['спортивн', 'jogls', 'лосін', 'легінс'], 'odyag', 'sportyvni-shtany'),
    (['сукн', 'плаття'], 'odyag', 'sukni'),
    (['спідниц'], 'odyag', 'spidnytsi'),
    (['блуз'], 'odyag', 'bluzy'),
    (['піжам', 'халат'], 'odyag', 'pizhamy'),
    (['білизн', 'труси', 'бюстгальтер'], 'odyag', 'bilyzna'),
    (['купальник'], 'odyag', 'kupalniky'),
    (['костюм'], 'odyag', 'kostyumy'),
    (['комбінезон'], 'odyag', 'kombinezony'),
    (['дитяч', 'kids', 'baby'], 'odyag', 'dytiachyi-odyag'),
    (['робочий'], 'odyag', 'inshe-odyag'),
]

def detect_category(name):
    name_lower = name.lower()
    for keywords, cat, subcat in CATEGORY_RULES:
        for kw in keywords:
            if kw.lower() in name_lower:
                return cat, subcat
    return 'odyag', 'inshe-odyag'

# ─── Load unit type mapping ─────────────────────────────────────────────────

def load_sht_articles():
    path = os.path.join(ROOT, "Одиниці виміру - €ШТ.  – копія.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    articles = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0]:
            articles.add(str(row[0]).strip())
    return articles

# ─── Parse products ─────────────────────────────────────────────────────────

def parse_products(sht_articles):
    path = os.path.join(ROOT, "Прайс (список товарів з посиланням 25.03.26 ).xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Аркуш1"]

    products = []
    seen_slugs = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        article, name_field, price, sale_price, qty, *_ = (list(row) + [None]*8)[:8]
        if not article or not name_field:
            continue

        article = str(article).strip()
        name, code, video_url, avg_weight = parse_name_field(name_field)

        if not name:
            continue

        quality = parse_quality(name)
        season = parse_season(name)
        cat_slug, subcat_slug = detect_category(name)
        price_unit = "piece" if article in sht_articles else "kg"

        slug = slugify(name)
        # Ensure unique slug
        if slug in seen_slugs:
            slug = f"{slug}-{article.lower().replace(' ', '-')}"
        if slug in seen_slugs:
            slug = f"{slug}-{code}" if code else f"{slug}-dup"
        seen_slugs.add(slug)

        product = {
            "articleCode": article,
            "code1C": code if code else None,
            "name": name,
            "slug": slug,
            "categorySlug": cat_slug,
            "subcategorySlug": subcat_slug,
            "quality": quality,
            "season": season,
            "country": "",
            "priceUnit": price_unit,
            "averageWeight": avg_weight,
            "videoUrl": video_url,
            "priceEur": float(price) if price else None,
            "salePriceEur": float(sale_price) if sale_price else None,
            "quantity": int(qty) if qty else None,
            "inStock": True,
        }
        products.append(product)

    return products

# ─── Parse lots ──────────────────────────────────────────────────────────────

def parse_lots():
    path = os.path.join(ROOT, "Список конкретних мішків 30.03.26.xlsx")
    wb = openpyxl.load_workbook(path)
    ws = wb["Аркуш1"]

    lots = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = (list(row) + [None]*10)[:10]
        article, name, weight, barcode, video_url, description, reserved, sale_val, price_val, rate = vals

        if not article and not name:
            continue

        article = str(article).strip() if article else ""
        barcode = str(barcode).strip() if barcode else ""
        weight = float(weight) if weight else 0

        # Determine status
        status = "free"
        if reserved and str(reserved).strip():
            status = "reserved"
        elif sale_val and str(sale_val).strip():
            status = "on_sale"

        # Parse price
        price_eur = 0
        if sale_val and str(sale_val).strip():
            try: price_eur = float(sale_val)
            except: pass
        if price_eur == 0 and price_val:
            try: price_eur = float(price_val)
            except: pass

        # Parse quantity from description
        qty = 1
        if description:
            qty_match = re.search(r'Кількість одиниць:\s*(\d+)', str(description))
            if qty_match:
                qty = int(qty_match.group(1))

        exchange_rate = float(rate) if rate else 50.9

        lot = {
            "articleCode": article,
            "productName": str(name).strip() if name else "",
            "barcode": barcode,
            "weight": weight,
            "quantity": qty,
            "status": status,
            "priceEur": round(price_eur / exchange_rate, 2) if price_eur > 100 else price_eur,
            "priceUah": price_eur if price_eur > 100 else round(price_eur * exchange_rate, 2),
            "videoUrl": str(video_url).strip() if video_url else None,
            "description": str(description).strip().replace('¶', '\n') if description else "",
            "reservedBy": str(reserved).strip() if reserved else None,
            "exchangeRate": exchange_rate,
        }
        lots.append(lot)

    return lots

# ─── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading unit type mapping...")
    sht_articles = load_sht_articles()
    print(f"  {len(sht_articles)} articles priced per unit (SHT)")

    print("Parsing products...")
    products = parse_products(sht_articles)
    print(f"  {len(products)} products parsed")

    print("Parsing lots...")
    lots = parse_lots()
    print(f"  {len(lots)} lots parsed")

    # Stats
    cats = {}
    for p in products:
        key = p["categorySlug"]
        cats[key] = cats.get(key, 0) + 1
    print("\nProducts by category:")
    for cat, count in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")

    lot_stats = {"free": 0, "reserved": 0, "on_sale": 0}
    for l in lots:
        lot_stats[l["status"]] += 1
    print(f"\nLots: {lot_stats}")

    # Write JSON
    os.makedirs(DATA_OUT, exist_ok=True)

    with open(os.path.join(DATA_OUT, "products.json"), "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"\nWrote {DATA_OUT}/products.json")

    with open(os.path.join(DATA_OUT, "lots.json"), "w", encoding="utf-8") as f:
        json.dump(lots, f, ensure_ascii=False, indent=2)
    print(f"Wrote {DATA_OUT}/lots.json")
